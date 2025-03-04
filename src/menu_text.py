import openpyxl

import settings
from dir_path import get_lower_directory_path


def get_tr():
    return tr


def _get_translate_from_file(language, table_name):
    file_path = get_lower_directory_path("") + table_name
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    try:
        key_index = headers.index("Dictionary key")
    except ValueError:
        raise ValueError("'Dictionary key' not found in Excel")

    try:
        lang_index = headers.index(language)
    except ValueError:
        raise ValueError(f"Row not found for language: {language}")

    translate_dictionary = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        dict_key = row[key_index]
        translation = row[lang_index]
        translate_dictionary[dict_key] = translation

    return translate_dictionary


def set_language():
    tr = _get_translate_from_file(
        settings.get_settings()["interface_language"], "translation.xlsx"
    )


tr = _get_translate_from_file(
    settings.get_settings()["interface_language"], "translation.xlsx"
)
